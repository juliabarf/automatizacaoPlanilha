import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import math
from decimal import Decimal, getcontext
from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill
import numpy as np
import os

# -----------------------------
def permeabilit(phi, fzi):
    phi = np.clip(phi, 1e-6, 0.99)
    phi_e = phi / (1 - phi)
    k = phi * ((fzi * phi_e) / 0.0314) ** 2
    return k


class AutomatizacaoPlanilha:
    def __init__(self, df, nomeTabela):
        self._df = df.copy()
        self.nomeTabela = nomeTabela

        self._df.columns = self._df.columns.str.strip()

        convert_dic = {'Prof. (m)': float}
        self._df = self._df.astype(convert_dic)
        self._profundidade = self._df['Prof. (m)']
        self._porosidade = self._df['Porosidade (%)']
        self._permeabilidade = self._df['Perm Abs Longitud (mD)']

    def porosidade(self):
        return list(self._porosidade)

    getcontext().prec = 28

    def porosidadeDec(self):
        return [Decimal(str(p)) / Decimal("100") for p in self._porosidade]

    def rqi(self):
        resultado = []
        for i in range(len(self._df)):
            try:
                permeabilidade = Decimal(str(self._permeabilidade[i]))
                porosidade_dec = self.porosidadeDec()[i]
                if permeabilidade == 0 or porosidade_dec == 0:
                    resultado.append(Decimal("0"))
                else:
                    rqi = Decimal("0.0314") * (permeabilidade / porosidade_dec).sqrt()
                    resultado.append(rqi)
            except Exception:
                resultado.append(Decimal("0"))
        return resultado

    def phi(self):
        resultado = []
        for i in range(len(self._df)):
            try:
                porosidade = self.porosidadeDec()[i]
                if porosidade <= 0 or porosidade >= 1:
                    resultado.append(Decimal("0"))
                else:
                    phi = porosidade / (Decimal("1") - porosidade)
                    resultado.append(phi)
            except Exception:
                resultado.append(Decimal("0"))
        return resultado

    def fzi(self):
        phi = self.phi()
        rqi = self.rqi()
        return [(r / p if p != 0 else Decimal("0")) for r, p in zip(rqi, phi)]

    def ghe(self):
        fzi = self.fzi()
        resultado = []
        for valor in fzi:
            try:
                if valor < 0.0938:
                    resultado.append(0)
                elif valor < 0.1875:
                    resultado.append(1)
                elif valor < 0.3750:
                    resultado.append(2)
                elif valor < 0.7500:
                    resultado.append(3)
                elif valor < 1.5000:
                    resultado.append(4)
                elif valor < 3.0000:
                    resultado.append(5)
                elif valor < 6.0000:
                    resultado.append(6)
                elif valor < 12.0000:
                    resultado.append(7)
                elif valor < 24.0000:
                    resultado.append(8)
                elif valor < 48.0000:
                    resultado.append(9)
                else:
                    resultado.append(10)
            except:
                resultado.append(0)
        return resultado

    def criaPlanilha(self):
        colunas = {
            'Profundidade': self._profundidade,
            'Porosity (%)': self.porosidade(),
            'Porosity Decimal': self.porosidadeDec(),
            'Permeability (mD)': self._permeabilidade,
            'RQI': self.rqi(),
            'PHI(Z)': self.phi(),
            'FZI': self.fzi(),
            'GHE': self.ghe()
        }
        dfColunas = pd.DataFrame(colunas).fillna(0)

        file_path = self.nomeTabela + 'Alterada.xlsx'

        if os.path.exists(file_path):
            workbook = load_workbook(file_path)
        else:
            workbook = Workbook()
            if "Sheet" in workbook.sheetnames:
                workbook.remove(workbook["Sheet"])

        if "Planilha1" in workbook.sheetnames:
            sheet1 = workbook["Planilha1"]
            for row in sheet1.iter_rows():
                for cell in row:
                    cell.value = None
        else:
            sheet1 = workbook.create_sheet("Planilha1")

        headers = list(dfColunas.columns)
        for col_num, header in enumerate(headers):
            cell = sheet1.cell(row=1, column=col_num + 1, value=header)
            cell.alignment = Alignment(horizontal='center', vertical='center')

            if header in ['FZI', 'RQI', 'PHI(Z)', 'GHE']:
                cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
            elif header in ['Profundidade', 'Porosity (%)', 'Porosity Decimal', 'Permeability (mD)']:
                cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

        for row_num, row in enumerate(dfColunas.itertuples(index=False), start=2):
            for col_num, value in enumerate(row):
                cell = sheet1.cell(row=row_num, column=col_num + 1,
                                   value=float(value) if isinstance(value, Decimal) else value)
                cell.alignment = Alignment(horizontal='center', vertical='center')

                if headers[col_num] in ['Porosity Decimal', 'Profundidade', 'Permeability (mD)', 'Porosity (%)', 'PHI(Z)']:
                    cell.number_format = '0.000'
                elif headers[col_num] in ['RQI', 'FZI']:
                    cell.number_format = '0.############'

        porosidade_dec = [float(p) for p in dfColunas['Porosity Decimal']]
        permeability = [float(p) for p in dfColunas['Permeability (mD)']]
        ghe = [int(g) for g in dfColunas['GHE']]

        if "Dados Grafico" in workbook.sheetnames:
            sheet_dados_grafico = workbook["Dados Grafico"]
            for row in sheet_dados_grafico.iter_rows():
                for cell in row:
                    cell.value = None
        else:
            sheet_dados_grafico = workbook.create_sheet("Dados Grafico")

        fzi_values = [48, 24, 12, 6, 3, 1.5, 0.75, 0.375, 0.1875, 0.0938]
        ghe_labels = list(range(10, 0, -1))
        phi = np.linspace(0.01, 0.5, 300)

        # FAIXAS
        data = []
        for i in range(len(fzi_values)):
            k = permeabilit(phi, fzi_values[i])
            data.append(k)

        col_faixas_porosidade = 1
        sheet_dados_grafico.cell(row=1, column=col_faixas_porosidade, value="Porosity (Faixas)")

        col_faixas_start = 2

        for i, label in enumerate(ghe_labels):
            sheet_dados_grafico.cell(row=1, column=col_faixas_start + i, value=f"Faixa GHE_{label}")

        for row, p in enumerate(phi, start=2):
            sheet_dados_grafico.cell(row=row, column=col_faixas_porosidade, value=p)
            for col, k_array in enumerate(data, start=col_faixas_start):
                sheet_dados_grafico.cell(row=row, column=col, value=k_array[row - 2])

        next_col = col_faixas_start + len(ghe_labels)
        ghe_columns = {}

        for ghe_valor in range(11):
            col_porosity = next_col
            col_perm = next_col + 1
            ghe_columns[ghe_valor] = (col_porosity, col_perm)

            sheet_dados_grafico.cell(row=1, column=col_porosity, value=f"Porosity Exp GHE {ghe_valor}")
            sheet_dados_grafico.cell(row=1, column=col_perm, value=f"Perm Exp GHE {ghe_valor}")

            current_row = 2

            for idx in range(len(porosidade_dec)):
                if ghe[idx] == ghe_valor:
                    sheet_dados_grafico.cell(row=current_row, column=col_porosity, value=porosidade_dec[idx])
                    sheet_dados_grafico.cell(row=current_row, column=col_perm, value=permeability[idx])
                    current_row += 1

            next_col += 2

        total_cols_grafico = next_col - 1
        for col in range(1, total_cols_grafico + 1):
            sheet_dados_grafico.column_dimensions[get_column_letter(col)].width = 20


        # ============================================================
        #       <<<  LINHAS X E Y  >>>
        # ============================================================

        linha_ref_col_x = next_col + 1
        linha_ref_col_y = next_col + 3

        sheet_dados_grafico.cell(row=1, column=linha_ref_col_x,     value="LinhaX_x")
        sheet_dados_grafico.cell(row=1, column=linha_ref_col_x + 1, value="LinhaX_y")
        sheet_dados_grafico.cell(row=1, column=linha_ref_col_y,     value="LinhaY_x")
        sheet_dados_grafico.cell(row=1, column=linha_ref_col_y + 1, value="LinhaY_y")

        # Linha vertical X = 0.1
        x_const = 0.1
        sheet_dados_grafico.cell(row=2,   column=linha_ref_col_x, value=x_const)
        sheet_dados_grafico.cell(row=301, column=linha_ref_col_x, value=x_const)
        sheet_dados_grafico.cell(row=2,   column=linha_ref_col_x + 1, value=1e-3)
        sheet_dados_grafico.cell(row=301, column=linha_ref_col_x + 1, value=1e6)

        # Linha horizontal Y = 1.0
        y_const = 1.0
        sheet_dados_grafico.cell(row=2,   column=linha_ref_col_y, value=0.0)
        sheet_dados_grafico.cell(row=301, column=linha_ref_col_y, value=0.6)
        sheet_dados_grafico.cell(row=2,   column=linha_ref_col_y + 1, value=y_const)
        sheet_dados_grafico.cell(row=301, column=linha_ref_col_y + 1, value=y_const)

        next_col = linha_ref_col_y + 2


        # ============================================================

        chart = ScatterChart()
        chart.title = "Global Hydraulic Elements"  # Alterado para corresponder à imagem
        chart.x_axis.title = "Porosity (decimal)"
        chart.y_axis.title = "Permeability (mD)"
        chart.legend.position = "r"
        chart.width = 20
        chart.height = 15

        # --- Eixo Y (Permeabilidade - Logarítmico) ---
        chart.y_axis.scaling.logBase = 10
        chart.y_axis.scaling.min = 0.001  # Mínimo de 0.001 (10^-3)
        chart.y_axis.scaling.max = 10000000  # Máximo de 10,000,000 (10^7)
        # As unidades e a exibição logarítmica são frequentemente controladas
        # pelo próprio Excel com base na escala logBase.
        # Manteremos as configurações mais importantes aqui:

        # --- Eixo X (Porosidade - Decimal) ---
        chart.x_axis.scaling.min = 0.0
        chart.x_axis.scaling.max = 0.6
        chart.x_axis.majorUnit = 0.1  # Marcações principais a cada 0.1 (0.0, 0.1, 0.2...)

        # Ajustes visuais
        chart.x_axis.majorGridlines = None
        chart.y_axis.majorGridlines = None

        chart.x_axis.scaling.min = -0.02  # Mantém o ajuste para começar um pouco antes de 0

        colors = [
            "FF0000", "FF4500", "FFA500", "FFD700", "ADFF2F",
            "00FA9A", "00CED1", "1E90FF", "8A2BE2", "FF69B4"
        ]

        col_faixas_y_start = col_faixas_start
        col_faixas_x = col_faixas_porosidade

        for i in range(len(fzi_values)):
            xvalues = Reference(sheet_dados_grafico, min_col=col_faixas_x, min_row=2, max_row=301)
            yvalues = Reference(sheet_dados_grafico, min_col=col_faixas_y_start + i, min_row=2, max_row=301)
            series = Series(yvalues, xvalues, title=f"GHE {ghe_labels[i]}")
            series.graphicalProperties.line.solidFill = colors[i]
            chart.append(series)

        max_row_exp = len(porosidade_dec) + 1

        for ghe_valor in range(11):
            col_porosity, col_perm = ghe_columns[ghe_valor]
            xvalues_exp = Reference(sheet_dados_grafico, min_col=col_porosity, min_row=2, max_row=max_row_exp)
            yvalues_exp = Reference(sheet_dados_grafico, min_col=col_perm, min_row=2, max_row=max_row_exp)

            series_exp = Series(yvalues_exp, xvalues_exp, title=f"GHE {ghe_valor}")
            series_exp.marker.symbol = "circle"
            series_exp.marker.size = 6
            series_exp.graphicalProperties.line.noFill = True

            if ghe_valor == 0:
                series_exp.marker.graphicalProperties.solidFill = "000000"
            else:
                idx = 10 - ghe_valor
                if idx > 0 and idx <= len(colors):
                    series_exp.marker.graphicalProperties.solidFill = colors[idx - 1]
                else:
                    series_exp.marker.graphicalProperties.solidFill = "000000"

            chart.append(series_exp)
            chart.series[-1].graphicalProperties.noFill = True


        # ------------------------------------------------------------
        #    ADICIONA AS LINHAS X E Y AO GRÁFICO
        # ------------------------------------------------------------

        xv_x = Reference(sheet_dados_grafico, min_col=linha_ref_col_x, max_col=linha_ref_col_x, min_row=2, max_row=301)
        xv_y = Reference(sheet_dados_grafico, min_col=linha_ref_col_x + 1, max_col=linha_ref_col_x + 1, min_row=2, max_row=301)

        linha_vertical = Series(xv_y, xv_x, title="Linha X")
        linha_vertical.graphicalProperties.line.solidFill = "000000"
        linha_vertical.graphicalProperties.line.width = 30000
        linha_vertical.marker = None
        chart.append(linha_vertical)

        yh_x = Reference(sheet_dados_grafico, min_col=linha_ref_col_y, max_col=linha_ref_col_y, min_row=2, max_row=301)
        yh_y = Reference(sheet_dados_grafico, min_col=linha_ref_col_y + 1, max_col=linha_ref_col_y + 1, min_row=2, max_row=301)

        linha_horizontal = Series(yh_y, yh_x, title="Linha Y")
        linha_horizontal.graphicalProperties.line.solidFill = "000000"
        linha_horizontal.graphicalProperties.line.width = 30000
        linha_horizontal.marker = None
        chart.append(linha_horizontal)

        # ------------------------------------------------------------

        sheet1.add_chart(chart, "J2")
        workbook.save(file_path)


class Aplicativo:
    def __init__(self, master=None):
        def selecionar_arquivo():
            arquivo = filedialog.askopenfilename(
                title="Selecione o arquivo .xlsx",
                filetypes=[("Planilhas Excel", "*.xlsx")]
            )
            print(arquivo)

            if arquivo:
                preview = pd.read_excel(arquivo, header=None)
                colunas_necessarias = ['Prof. (m)', 'Porosidade (%)', 'Perm Abs Longitud (mD)']
                header_row = None

                for i, row in preview.iterrows():
                    valores_validos = [str(v).strip() for v in row.values if pd.notna(v)]
                    if all(col in valores_validos for col in colunas_necessarias):
                        header_row = i
                        break

                if header_row is not None:
                    df = pd.read_excel(arquivo, header=header_row)
                    df.columns = df.columns.str.strip()

                    if all(col in df.columns for col in colunas_necessarias):
                        nomeTabela = arquivo.split('.')[0]
                        AutomatizacaoPlanilha(df, nomeTabela).criaPlanilha()
                        messagebox.showinfo('Sucesso', 'Sua planilha foi criada com sucesso!')
                    else:
                        cols_faltando = [col for col in colunas_necessarias if col not in df.columns]
                        messagebox.showerror('Erro', f'A planilha está faltando as colunas: {", ".join(cols_faltando)}')
                else:
                    messagebox.showerror('Erro', 'Não foi possível encontrar as colunas esperadas na planilha.')

        self.primeiroContainer = tk.Frame(master, pady=10, padx=100)
        self.primeiroContainer.pack()

        self.segundoContainer = tk.Frame(master, pady=10)
        self.segundoContainer.pack()

        self.terceiroContainer = tk.Frame(master, pady=5)
        self.terceiroContainer.pack()

        self.quartoContainer = tk.Frame(master, pady=10)
        self.quartoContainer.pack()

        self.titulo = tk.Label(
            self.primeiroContainer,
            text='Antes de selecionar o arquivo, observe\nse a planilha contém as seguintes colunas:\n\nProf. (m)\nPorosidade (%)\nPerm Abs Longitud (m)')

        self.titulo.pack()

        self.btnArquivo = tk.Button(self.segundoContainer, text='Selecionar arquivo', width=25,
                                    command=selecionar_arquivo)
        self.btnArquivo.pack()


root = tk.Tk()
root.title('Planilhas Petrofisica Lagesed')
app = Aplicativo(root)
root.mainloop()
