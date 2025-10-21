import openpyxl
from openpyxl.chart import BarChart, Reference
import numpy as np
import pandas as pd
import xlsxwriter
import os

def permeabilit(phi, fzi):
    phi = np.clip(phi, 1e-6, 0.99)
    phi_e = phi / (1 - phi)
    k = phi * ((fzi * phi_e) / 0.0314) ** 2
    return k



def principal(file_path):
    fzi_values = [48, 24, 12, 6, 3, 1.5, 0.75, 0.375, 0.1875, 0.0938]
    ghe_labels = list(range(10, 0, -1))
    phi = np.linspace(0.01, 0.5, 300)

    # Carrega o arquivo existente
    wb = openpyxl.load_workbook(file_path)

    # Calcula os dados
    data = []
    for fzi in fzi_values:
        k = permeabilit(phi, fzi)
        data.append(k)

    # Cria o DataFrame
    faixa_df = pd.DataFrame({"Porosity": phi})
    for i, k in enumerate(data):
        faixa_df[f"GHE_{ghe_labels[i]}"] = k

    # Remove a aba "Faixas" se já existir
    if "Faixas" in wb.sheetnames:
        del wb["Faixas"]

    # Cria uma nova aba "Faixas"
    ws_faixas = wb.create_sheet("Faixas")

    # Escreve o DataFrame na aba "Faixas"
    for r_idx, row in enumerate(faixa_df.itertuples(index=False), start=2):
        if r_idx == 2:  # cabeçalho
            for c_idx, col_name in enumerate(faixa_df.columns, start=1):
                ws_faixas.cell(row=1, column=c_idx, value=col_name)
        for c_idx, value in enumerate(row, start=1):
            ws_faixas.cell(row=r_idx, column=c_idx, value=value)



    wb.save(file_path)
    return ws_faixas


